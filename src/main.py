# src/main.py

import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from data_quality import data_quality_checks
from textual_report import generate_textual_report


from sklearn.linear_model import LogisticRegression
from sklearn.neural_network import MLPClassifier
import statsmodels.api as sm

# -------------------
# Custom Modules
# -------------------
from financial_ratios import calculate_all_financial_ratios
from market_ratios import add_market_ratios  # Phase 6: Market/Valuation Ratios

# -------------------
# Paths
# -------------------
input_file = "../data/financial_data.xlsx"
output_file = "../outputs/financial_analysis_results.xlsx"
os.makedirs(os.path.dirname(output_file), exist_ok=True)

# -------------------
# Load Data
# -------------------
def load_data(file_path):
    if os.path.exists(file_path):
        df = pd.read_excel(file_path)
        print("Data loaded from file.")
    else:
        # Sample data if file not found
        df = pd.DataFrame({
            'Company': ['A', 'B', 'C'],
            'CurrentAssets': [100000, 150000, 120000],
            'CurrentLiabilities': [50000, 70000, 60000],
            'Inventory': [20000, 30000, 25000],
            'Cash': [30000, 40000, 35000],
            'Revenue': [200000, 260000, 240000],
            'COGS': [120000, 150000, 140000],
            'OperatingIncome': [30000, 35000, 32000],
            'NetIncome': [20000, 25000, 23000],
            'TotalAssets': [180000, 220000, 200000],
            'Equity': [80000, 100000, 90000],
            'TotalLiabilities': [100000, 120000, 110000],
            'AccountsReceivable': [30000, 40000, 35000],
            'FCF': [5000, 7000, 6000],
            'CostOfDebt': [0.05, 0.06, 0.055],
            'TaxRate': [0.25, 0.25, 0.25],
            'Beta': [1.0, 0.9, 0.95],
            'Excess_Return': [0.12, 0.09, 0.10],
            'Market_Excess': [0.10, 0.10, 0.10],
            'SMB': [0.02, 0.02, 0.02],
            'HML': [0.01, 0.01, 0.01],
            'Z_Score': [3.0, 2.2, 2.8],
            'Distress': [0, 1, 0],
            # Market/Phase6 data
            'MarketPrice': [50, 70, 60],
            'EPS': [5, 7, 6],
            'BookValuePerShare': [40, 60, 50],
            'MarketCap': [500000, 700000, 600000],
            'EBITDA': [35000, 42000, 38000],
            'DividendPerShare': [1.5, 2, 1.8],
            'SharesOutstanding': [10000, 10000, 10000]
        })
        print("Sample data created.")
    return df

def main():
    df = load_data(input_file)

    # Phase 0: Data Quality Control
    df = data_quality_checks(df)

    # Phase 1: Financial Ratios
    df = calculate_all_financial_ratios(df)

    # Phase 6: Market Ratios
    df = add_market_ratios(df)

    # Phase 2: Valuation
    df = capm_expected_return(df)
    df = calculate_wacc(df)
    df = dcf_valuation(df)

    # Phase 3: Asset Pricing & Distress
    df = fama_french_regression(df)
    df = ml_financial_distress(df)

    save_results_with_charts(df)

# -------------------
# CAPM
# -------------------
def capm_expected_return(df, risk_free=0.03, market_return=0.10):
    if 'Beta' in df.columns:
        df['Expected_Return'] = risk_free + df['Beta'] * (market_return - risk_free)
    return df

# -------------------
# WACC
# -------------------
def calculate_wacc(df):
    required = ['Equity', 'TotalLiabilities', 'CostOfDebt', 'TaxRate', 'Expected_Return']
    if all(col in df.columns for col in required):
        df['V'] = df['Equity'] + df['TotalLiabilities']
        df['E_V'] = df['Equity'] / df['V']
        df['D_V'] = df['TotalLiabilities'] / df['V']
        df['WACC'] = df['E_V'] * df['Expected_Return'] + df['D_V'] * df['CostOfDebt'] * (1 - df['TaxRate'])
    return df

# -------------------
# DCF Valuation
# -------------------
def dcf_valuation(df):
    if all(col in df.columns for col in ['FCF', 'WACC']):
        df['DCF_Value'] = df['FCF'] / df['WACC']
    return df

# -------------------
# Fama-French Regression
# -------------------
def fama_french_regression(df):
    required = ['Excess_Return', 'Market_Excess', 'SMB', 'HML']
    if all(col in df.columns for col in required):
        X = df[['Market_Excess', 'SMB', 'HML']]
        X = sm.add_constant(X)
        y = df['Excess_Return']
        model = sm.OLS(y, X).fit()
        df['Alpha'] = model.params.get('const', np.nan)
        df['Beta_Market'] = model.params.get('Market_Excess', np.nan)
        df['Beta_SMB'] = model.params.get('SMB', np.nan)
        df['Beta_HML'] = model.params.get('HML', np.nan)
    return df

# -------------------
# Financial Distress ML
# -------------------
def ml_financial_distress(df):
    if all(col in df.columns for col in ['Z_Score', 'Distress']):
        X = df[['Z_Score']]
        y = df['Distress']
        if len(np.unique(y)) > 1:
            logreg = LogisticRegression(max_iter=500)
            logreg.fit(X, y)
            df['Distress_LogReg_Pred'] = logreg.predict(X)

            nn = MLPClassifier(hidden_layer_sizes=(5,5), max_iter=500)
            nn.fit(X, y)
            df['Distress_NN_Pred'] = nn.predict(X)
    return df

# -------------------
# Save Results & Charts
# -------------------
def save_results_with_charts(df):
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Summary', index=False)

    wb = load_workbook(output_file)
    ws = wb.create_sheet(title='Charts')

    numeric_cols = df.select_dtypes(include=[np.number]).columns

    for i, col in enumerate(numeric_cols):
        plt.figure(figsize=(6,4))
        sns.barplot(x='Company', y=col, data=df)
        plt.title(col)
        buf = BytesIO()
        plt.savefig(buf, format='png')
        plt.close()
        buf.seek(0)
        img = Image(buf)
        ws.add_image(img, f"A{1 + i*15}")

    wb.save(output_file)
    print(f"Final processed data and charts saved to {output_file}")

# -------------------
# Main Pipeline
# -------------------
def main():
    df = load_data(input_file)

    # Phase 1: Financial Ratios
    df = calculate_all_financial_ratios(df)

    # Phase 6: Market Ratios
    df = add_market_ratios(df)

    # Phase 2: Valuation
    df = capm_expected_return(df)
    df = calculate_wacc(df)
    df = dcf_valuation(df)

    # Phase 3: Asset Pricing & Distress
    df = fama_french_regression(df)
    df = ml_financial_distress(df)

    save_results_with_charts(df)
    generate_textual_report(df)

if __name__ == "__main__":
    main()
